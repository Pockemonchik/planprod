import openpyxl 

# ������� ��� ��������� �������������� ������ ���� ������ ���������� �� xls ����� � ��� �������
def takeXls(nameDoc,namePrepod):
    wb = openpyxl.load_workbook(filename =  '/home/andrey/Documents/Vazhno_sho_pizdec/mvdproject/mvd/plan/' + nameDoc + '.xlsx', data_only=True)
    
    needColumnFirst = ['O','Q','S','U','W','X','Y','AA','AC','AE','AG','AI','AK','AM','AO','AQ','AS','AU','AW','AY','BA','BC','BE','BF','BG' ]
    needColumnSecond = ['BZ','CA','CB','CD','CF','CH','CI','CJ','CL','CN','CP','CR','CT','CV','CX','CZ','DB','DD','DF','DH','DJ','DL','DN','DP','DQ','DR']
    needColumnThird = ['EK','EM','EO','EQ','ES','ET','EU','EW','EY','FA','FC','FE','FG','FI','FK','FM','FO','FQ','FS','FU','FW','FY','GA','GB','GC']
    listFirst = []
    listSecond = []
    listResult = []
    
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    prepodRow = 1
    #������� ������ � ��������
    for column in ws.iter_rows(min_col=2, min_row=1, max_col=2, max_row=400, values_only=True):
        if str(column[0]).lower() == str(namePrepod).lower():
            break
        prepodRow += 1
        
    #�������� �������� 1 ������� � ����
    
    
    for takeStr in range(1,14):
        listFirst.append([])
        firstBigSell = str(ws['B'+str(prepodRow+takeStr)].value) +' '+ str(ws['E'+str(prepodRow+takeStr)].value) +' '+ str(ws['F'+str(prepodRow+takeStr)].value)
        listFirst[takeStr-1].append(firstBigSell)
        for takeColumn in needColumnFirst:
            cell = takeColumn+str(prepodRow+takeStr)
            listFirst[takeStr-1].append(ws[cell].value)
    listFirst.append([])
    for takeColumn in needColumnFirst:
            cell = takeColumn+str(prepodRow)
            listFirst[takeStr].append(ws[cell].value)
            
            
            
    #�������� �������� 2 ������� � ����
    for takeStr in range(1,13):
        listSecond.append([])
        firstBigSell = str(ws['BM'+str(prepodRow+takeStr)].value) +' '+ str(ws['BP'+str(prepodRow+takeStr)].value) +' '+ str(ws['BQ'+str(prepodRow+takeStr)].value)
        listSecond[takeStr-1].append(firstBigSell)
        for takeColumn in needColumnSecond:
            cell = takeColumn+str(prepodRow+takeStr-1)
            listSecond[takeStr-1].append(ws[cell].value)
    listSecond.append([])
    for takeColumn in needColumnSecond:
            cell = takeColumn+str(prepodRow)
            listSecond[takeStr].append(ws[cell].value)
    listSecond.append([])
    #�������� �������� 3 ������� � ����
    for takeColumn in needColumnThird:
            cell = takeColumn+str(prepodRow)
            listSecond[takeStr+1].append(ws[cell].value)
    
    listResult.append(listFirst)
    listResult.append(listSecond)
    return(listResult)